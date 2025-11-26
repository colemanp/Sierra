"""MacroFactor nutrition XLSX importer"""
from pathlib import Path
from typing import Any, Dict, Iterator, Optional
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from .base import BaseImporter


class MacroFactorImporter(BaseImporter):
    """Import MacroFactor nutrition XLSX exports

    MacroFactor exports contain daily nutrition summaries with:
    - Date, weight, calories, macros, micros
    - Individual food entries with timestamps
    """

    SOURCE_NAME = "macrofactor"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if load_workbook is None:
            raise ImportError("openpyxl is required for MacroFactor imports. Install with: pip install openpyxl")

    def _parse_file(self, file_path: Path) -> Iterator[Dict[str, Any]]:
        """Parse MacroFactor XLSX file"""
        wb = load_workbook(file_path, read_only=True, data_only=True)

        # Try to find the main data sheet
        sheet = wb.active

        # Get headers from first row
        headers = []
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(row)]
            break

        if not headers:
            self.logger.warning("No headers found in XLSX file")
            return

        # Process data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue

            record = dict(zip(headers, row))
            yield record

        wb.close()

    def _process_record(self, record: Dict[str, Any]) -> str:
        """Process a single nutrition record"""
        # Find date column (various possible names)
        date_value = None
        for key in ["date", "day", "logged date", "entry date"]:
            if key in record and record[key]:
                date_value = record[key]
                break

        if not date_value:
            return "skipped"

        # Parse date
        date_iso = self._parse_date(date_value)
        if not date_iso:
            self.logger.warning(f"Skipping record with invalid date: {date_value}")
            return "skipped"

        # Check if this is a daily summary or food entry
        # Daily summaries typically have expenditure/target columns
        is_daily = any(k in record for k in ["expenditure", "target calories", "tdee"])

        if is_daily:
            return self._process_daily_record(date_iso, record)
        else:
            return self._process_food_entry(date_iso, record)

    def _process_daily_record(self, date_iso: str, record: Dict[str, Any]) -> str:
        """Process daily nutrition summary"""
        data = {
            "source_id": self.source_id,
            "date": date_iso,
            "expenditure_kcal": self._get_float(record, ["expenditure", "tdee", "energy expenditure"]),
            "calories_consumed_kcal": self._get_float(record, ["calories", "energy", "kcal", "calories consumed"]),
            "target_calories_kcal": self._get_float(record, ["target calories", "calorie target", "goal"]),
            "weight_lbs": self._get_weight_lbs(record),
            "trend_weight_lbs": self._get_float(record, ["trend weight", "smoothed weight"]),
            "protein_g": self._get_float(record, ["protein", "protein (g)"]),
            "fat_g": self._get_float(record, ["fat", "fat (g)", "total fat"]),
            "carbs_g": self._get_float(record, ["carbs", "carbohydrates", "carbs (g)"]),
            "fiber_g": self._get_float(record, ["fiber", "fibre", "dietary fiber"]),
            "alcohol_g": self._get_float(record, ["alcohol", "alcohol (g)"]),
            "target_protein_g": self._get_float(record, ["target protein", "protein target"]),
            "target_fat_g": self._get_float(record, ["target fat", "fat target"]),
            "target_carbs_g": self._get_float(record, ["target carbs", "carb target"]),
            "steps": self._get_int(record, ["steps", "step count"]),
        }

        key_fields = {
            "source_id": self.source_id,
            "date": date_iso,
        }

        return self._insert_with_conflict_check(
            "nutrition_daily",
            key_fields,
            data
        )

    def _process_food_entry(self, date_iso: str, record: Dict[str, Any]) -> str:
        """Process individual food entry"""
        food_name = self._get_string(record, ["food", "food name", "name", "item", "description"])
        if not food_name:
            return "skipped"

        time_value = self._get_string(record, ["time", "logged time", "meal time"])
        time_24h = self._parse_time(time_value) if time_value else None

        data = {
            "source_id": self.source_id,
            "date": date_iso,
            "time": time_24h,
            "food_name": food_name,
            "serving_size": self._get_string(record, ["serving", "serving size", "portion"]),
            "serving_qty": self._get_float(record, ["quantity", "servings", "amount"]),
            "serving_weight_g": self._get_float(record, ["weight", "grams", "weight (g)"]),
            "calories_kcal": self._get_float(record, ["calories", "energy", "kcal"]),
            "protein_g": self._get_float(record, ["protein", "protein (g)"]),
            "fat_g": self._get_float(record, ["fat", "fat (g)"]),
            "carbs_g": self._get_float(record, ["carbs", "carbohydrates"]),
            "fiber_g": self._get_float(record, ["fiber", "fibre"]),
        }

        key_fields = {
            "source_id": self.source_id,
            "date": date_iso,
            "time": time_24h,
            "food_name": food_name,
        }

        return self._insert_with_conflict_check(
            "nutrition_entries",
            key_fields,
            data
        )

    def _get_float(self, record: Dict, keys: list) -> Optional[float]:
        """Get float value from record, trying multiple possible keys"""
        for key in keys:
            if key in record and record[key] is not None:
                try:
                    return float(record[key])
                except (ValueError, TypeError):
                    continue
        return None

    def _get_int(self, record: Dict, keys: list) -> Optional[int]:
        """Get int value from record"""
        val = self._get_float(record, keys)
        return int(val) if val is not None else None

    def _get_string(self, record: Dict, keys: list) -> Optional[str]:
        """Get string value from record"""
        for key in keys:
            if key in record and record[key]:
                return str(record[key]).strip()
        return None

    def _get_weight_lbs(self, record: Dict) -> Optional[float]:
        """Get weight in lbs, converting from kg if needed"""
        # Try lbs first
        lbs = self._get_float(record, ["weight (lbs)", "weight_lbs", "weight lbs"])
        if lbs is not None:
            return lbs

        # Try kg and convert
        kg = self._get_float(record, ["weight", "weight (kg)", "weight_kg"])
        if kg is not None:
            return kg * 2.20462

        return None

    def _parse_date(self, value: Any) -> Optional[str]:
        """Parse date value to ISO format"""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, str):
            # Try common formats
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]:
                try:
                    dt = datetime.strptime(value, fmt)
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue
        return None

    def _parse_time(self, value: Any) -> Optional[str]:
        """Parse time value to HH:MM:SS format"""
        if isinstance(value, datetime):
            return value.strftime("%H:%M:%S")
        if isinstance(value, str):
            for fmt in ["%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"]:
                try:
                    dt = datetime.strptime(value.upper(), fmt)
                    return dt.strftime("%H:%M:%S")
                except ValueError:
                    continue
        return None

    def _log_insert(self, record: Dict[str, Any]) -> None:
        """Log insert"""
        date = record.get("date", "")
        self.logger.info(f"Inserted: nutrition {date}")

    def _log_skip(self, record: Dict[str, Any]) -> None:
        """Log skip"""
        date = record.get("date", "")
        self.logger.debug(f"Skipped: nutrition {date} (already exists)")
